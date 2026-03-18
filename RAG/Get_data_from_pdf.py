import os
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment


def extract_pdf_text(pdf_path):
    """
    提取PDF文字（修复单栏文字顺序、优化摘要/关键词段落结构）
    :param pdf_path: PDF文件路径
    :return: 按正确阅读顺序排列的文字
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            full_text = []
            for page in pdf.pages:
                words = page.extract_words()
                if not words:
                    continue

                # 1. 提取坐标信息，判断单/双栏
                x_coords = [w["x0"] for w in words]
                page_width = page.width
                split_x = page_width / 2
                total_words = len(words)
                left_count = sum(1 for x in x_coords if x < split_x)
                right_count = total_words - left_count
                is_two_column = (left_count / total_words > 0.2) and (right_count / total_words > 0.2)

                if is_two_column:
                    # 双栏处理：先左后右，每行按x排序
                    left_words = [w for w in words if w["x0"] < split_x]
                    right_words = [w for w in words if w["x0"] >= split_x]

                    # 排序：先按top（行），再按x0（列）
                    left_sorted = sorted(left_words, key=lambda w: (w["top"], w["x0"]))
                    right_sorted = sorted(right_words, key=lambda w: (w["top"], w["x0"]))

                    # 拼接左栏
                    left_text = ""
                    current_top = None
                    for w in left_sorted:
                        if current_top is None:
                            current_top = w["top"]
                        # 行内换行阈值：y差>8（适配PDF行距）
                        if abs(w["top"] - current_top) > 8:
                            left_text += "\n"
                            current_top = w["top"]
                        left_text += w["text"] + " "

                    # 拼接右栏
                    right_text = ""
                    current_top = None
                    for w in right_sorted:
                        if current_top is None:
                            current_top = w["top"]
                        if abs(w["top"] - current_top) > 8:
                            right_text += "\n"
                            current_top = w["top"]
                        right_text += w["text"] + " "

                    page_text = f"{left_text.strip()}\n\n{right_text.strip()}"

                else:
                    # ---------------------- 核心修复：单栏排序逻辑 ----------------------
                    # 先按top（行）排序，同一行内按x0（列）排序 → 保证同一行文字从左到右
                    words_sorted = sorted(words, key=lambda w: (w["top"], w["x0"]))

                    page_text = ""
                    current_top = None
                    # 优化换行阈值：行内换行（小阈值）vs 段落换行（大阈值）
                    for idx, w in enumerate(words_sorted):
                        if current_top is None:
                            current_top = w["top"]

                        # 1. 行内换行：y差>8 → 普通换行（行结束）
                        if abs(w["top"] - current_top) > 8:
                            page_text += "\n"
                            current_top = w["top"]
                        # 2. 段落换行：关键词/摘要/标题后加空行（识别特殊标记）
                        elif any(tag in w["text"] for tag in ["摘要：", "关键词：", "中图分类号：", "0引言"]):
                            page_text += "\n\n" + w["text"] + " "
                        # 3. 正常文字：拼接
                        else:
                            page_text += w["text"] + " "

                    # 清理多余空格和换行，修复摘要顺序
                    page_text = page_text.strip()
                    # 特殊修复：把"摘要："后的文字归位（针对你的文本格式）
                    if "摘 要：" in page_text:
                        abstract_idx = page_text.index("摘 要：")
                        # 把"摘 要："前的多余文字移到后面
                        before_abstract = page_text[:abstract_idx].strip()
                        abstract_content = page_text[abstract_idx:].strip()
                        if before_abstract and abstract_content:
                            page_text = f"{abstract_content}\n{before_abstract}"

                full_text.append(page_text)

            # 最终清理：合并多余换行，保证格式整洁
            final_text = "\n".join(full_text).replace("\n\n\n", "\n\n").strip()
            return final_text
    except Exception as e:
        return f"PDF提取失败：{str(e)}"


def pdf_to_xlsx(pdf_folder, output_xlsx_path):
    """
    遍历PDF文件夹，提取文字并生成XLSX文件
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "PDF文字提取结果"

    # 表头
    headers = ["title", "authors", "publish_year", "summary", "content", "dense_vector", "sparse_vector"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 遍历PDF文件
    pdf_files = [f for f in os.listdir(pdf_folder) if f.lower().endswith(".pdf")]
    if not pdf_files:
        print(f"⚠️ 未在 {pdf_folder} 中找到PDF文件！")
        return

    row_idx = 2
    for pdf_file in pdf_files:
        pdf_path = os.path.join(pdf_folder, pdf_file)
        print(f"正在处理：{pdf_file}")
        pdf_text = extract_pdf_text(pdf_path)

        # 写入Excel（仅填充content列）
        ws.cell(row=row_idx, column=5, value=pdf_text)
        ws.row_dimensions[row_idx].height = 100
        row_idx += 1

    # 调整列宽
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 80
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20

    wb.save(output_xlsx_path)
    print(f"\n✅ 处理完成！共处理 {len(pdf_files)} 个PDF，结果保存至：{output_xlsx_path}")


if __name__ == "__main__":
    # ====================== 请修改以下配置 ======================
    PDF_FOLDER = r"D:\AC\Agent_System\Data"  # 你的PDF文件夹路径
    OUTPUT_XLSX = r"D:\AC\Agent_System\Data\pdf_extract_result.xlsx"  # 输出Excel路径
    # ===========================================================
    pdf_to_xlsx(PDF_FOLDER, OUTPUT_XLSX)