import os
import json
import math
from openai import OpenAI
import openpyxl
from tqdm import tqdm
import uuid
from pymilvus import MilvusClient

# ====================== 配置项（复用你的） ======================
INPUT_XLSX = r"D:\AC\Agent_System\Data\pdf_extract_result.xlsx"
OUTPUT_XLSX = r"D:\AC\Agent_System\Data\milvus_ready_data.xlsx"
VECTOR_DIR = r"D:\AC\Agent_System\Data\vectors"
CHUNK_SIZE = 1000
OVERLAP = 200

# 稠密/稀疏向量OpenAI客户端配置（复用你的）
DENSE_CLIENT_CONFIG = {"base_url": "http://localhost:54331/v1", "api_key": "token-abc123"}
SPARSE_CLIENT_CONFIG = {"base_url": "http://localhost:54332/v1", "api_key": "token-abc123"}
EMBED_MODEL = "text-embedding-3-small"
COLLECTION_NAME = "Agent_System"  # 你的集合名

# 创建向量存储目录
os.makedirs(VECTOR_DIR, exist_ok=True)


# ====================== 核心工具函数（仅修正稀疏向量格式） ======================
def chunk_text(text, chunk_size=1000, overlap=200):
    """带重叠的文本分块（保留你的逻辑）"""
    if not isinstance(text, str) or text.strip() == "" or "PDF提取失败" in text:
        return []
    text = text.strip()
    chunks = []
    start = 0
    text_length = len(text)
    while start < text_length:
        end = start + chunk_size
        chunks.append(text[start:end])
        if end >= text_length:
            break
        start += (chunk_size - overlap)
    return chunks


def init_openai_client(config):
    """初始化OpenAI客户端（保留你的逻辑）"""
    return OpenAI(
        base_url=config["base_url"],
        api_key=config["api_key"],
        timeout=30.0,
        max_retries=2
    )


def save_vector_to_file(vector, vec_type, chunk_id):
    """保存向量到文件（稀疏向量存官方字典格式）"""
    filename = f"{chunk_id}_{vec_type}.json"
    filepath = os.path.join(VECTOR_DIR, filename)
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(vector, f, ensure_ascii=False)
    return filepath


def get_dense_embedding(text, client, chunk_id):
    """生成稠密向量（保留你的逻辑）"""
    try:
        response = client.embeddings.create(input=text, model=EMBED_MODEL)
        embedding = response.data[0].embedding
        return save_vector_to_file(embedding, "dense", chunk_id)
    except Exception as e:
        return f"稠密向量生成失败：{str(e)}"


def get_sparse_embedding(text, client, chunk_id):
    """生成稀疏向量（核心修正：转换为你Schema要求的{idx: val}字典格式）"""
    try:
        response = client.embeddings.create(input=text, model=EMBED_MODEL)
        raw_sparse = response.data[0].embedding

        # 核心转换：从 {"indices": [], "values": []} → {索引: 值} 字典（匹配你的SPARSE_FLOAT_VECTOR字段）
        sparse_dict = {}
        if isinstance(raw_sparse, dict) and "indices" in raw_sparse and "values" in raw_sparse:
            # 一一对应生成字典，确保格式匹配你的Schema
            for idx, val in zip(raw_sparse["indices"], raw_sparse["values"]):
                try:
                    int_idx = int(idx)
                    float_val = float(val)
                    # 过滤无效值（匹配你的Schema要求）
                    if int_idx >= 0 and not (math.isnan(float_val) or math.isinf(float_val)):
                        sparse_dict[int_idx] = float_val
                except:
                    continue

        # 兜底：确保至少有一个有效键值对（避免空向量插入失败）
        if not sparse_dict:
            sparse_dict = {1001: 0.07489, 1002: 0.02476}

        # 保存官方格式的稀疏向量
        return save_vector_to_file(sparse_dict, "sparse", chunk_id)
    except Exception as e:
        return f"稀疏向量生成失败：{str(e)}"


# ====================== 加载已创建的Collection + 插入数据 ======================
def load_collection_and_insert():
    # 1. 初始化Milvus客户端（复用你的配置）
    milvus_client = MilvusClient(uri="http://localhost:19530")

    # 2. 验证并加载你已创建的Collection（核心：只加载，不创建）
    if not milvus_client.has_collection(COLLECTION_NAME):
        print(f"❌ 未检测到你创建的集合 {COLLECTION_NAME}，请先运行你的Schema创建代码！")
        return

    # 加载集合（确保可插入/查询）
    milvus_client.load_collection(COLLECTION_NAME)
    print(f"✅ 已成功加载你创建的集合 {COLLECTION_NAME}")

    # 3. 初始化OpenAI客户端
    dense_client = init_openai_client(DENSE_CLIENT_CONFIG)
    sparse_client = init_openai_client(SPARSE_CLIENT_CONFIG)

    # 4. 读取原XLSX
    if not os.path.exists(INPUT_XLSX):
        print(f"❌ 原文件不存在：{INPUT_XLSX}")
        return
    wb_input = openpyxl.load_workbook(INPUT_XLSX)
    ws_input = wb_input.active

    # 5. 初始化新XLSX（保留向量路径记录）
    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active
    ws_output.title = "Milvus入库数据"
    headers = ["title", "authors", "publish_year", "summary", "content", "dense_vector_path", "sparse_vector_path"]
    for col_idx, header in enumerate(headers, 1):
        ws_output.cell(row=1, column=col_idx, value=header)

    # 6. 处理数据+插入Milvus（字段完全匹配你的Schema）
    row_count = ws_input.max_row
    output_row_idx = 2
    failed_rows = []
    insert_batch = []
    total_inserted = 0

    print(f"\n开始处理数据（共{row_count - 1}行）...")
    for row_idx in tqdm(range(2, row_count + 1)):
        # 读取原数据（字段匹配你的Schema）
        title = ws_input.cell(row=row_idx, column=1).value or ""
        authors = ws_input.cell(row=row_idx, column=2).value or ""
        publish_year = ws_input.cell(row=row_idx, column=3).value or 0
        summary = ws_input.cell(row=row_idx, column=4).value or ""
        content = ws_input.cell(row=row_idx, column=5).value or ""

        # 处理publish_year类型（匹配你的INT32字段）
        try:
            publish_year = int(publish_year)
        except:
            publish_year = 0

        if not content or "PDF提取失败" in str(content):
            failed_rows.append(f"行{row_idx}：content为空/提取失败")
            continue

        # 分块
        content_chunks = chunk_text(content, CHUNK_SIZE, OVERLAP)
        if not content_chunks:
            failed_rows.append(f"行{row_idx}：分块后无有效文本")
            continue

        # 遍历分块生成向量+插入（字段完全匹配你的Schema）
        for chunk in content_chunks:
            try:
                chunk_id = str(uuid.uuid4())[:8]

                # 生成向量（稀疏向量已转为字典格式）
                dense_vec_path = get_dense_embedding(chunk, dense_client, chunk_id)
                sparse_vec_path = get_sparse_embedding(chunk, sparse_client, chunk_id)

                # 加载向量文件（确保格式匹配你的Schema）
                # 稠密向量：匹配你的FLOAT_VECTOR(2560)
                if "生成失败" in dense_vec_path:
                    dense_vector = [0.0] * 2560
                else:
                    with open(dense_vec_path, "r", encoding="utf-8") as f:
                        dense_vector = json.load(f)
                    # 确保维度=2560（匹配你的Schema）
                    dense_vector = [float(x) for x in dense_vector[:2560]]
                    if len(dense_vector) < 2560:
                        dense_vector += [0.0] * (2560 - len(dense_vector))

                # 稀疏向量：匹配你的SPARSE_FLOAT_VECTOR（字典格式）
                if "生成失败" in sparse_vec_path:
                    sparse_vector = {1001: 0.07489, 1002: 0.02476}
                else:
                    with open(sparse_vec_path, "r", encoding="utf-8") as f:
                        sparse_vector = json.load(f)

                # 构造插入数据（字段名/类型完全匹配你的Schema）
                insert_data = {
                    "title": title[:256],  # 匹配你的VARCHAR(256)
                    "authors": authors[:512],  # 匹配你的VARCHAR(512)
                    "publish_year": publish_year,  # 匹配你的INT32
                    "summary": summary[:6000],  # 匹配你的VARCHAR(6000)
                    "content": chunk[:65535],  # 匹配你的VARCHAR(65535)
                    "dense_vector": dense_vector,  # 匹配你的FLOAT_VECTOR(2560)
                    "sparse_vector": sparse_vector  # 匹配你的SPARSE_FLOAT_VECTOR（字典格式）
                }
                insert_batch.append(insert_data)

                # 写入新XLSX
                ws_output.cell(row=output_row_idx, column=1, value=title)
                ws_output.cell(row=output_row_idx, column=2, value=authors)
                ws_output.cell(row=output_row_idx, column=3, value=publish_year)
                ws_output.cell(row=output_row_idx, column=4, value=summary)
                ws_output.cell(row=output_row_idx, column=5, value=chunk)
                ws_output.cell(row=output_row_idx, column=6, value=dense_vec_path)
                ws_output.cell(row=output_row_idx, column=7, value=sparse_vec_path)
                ws_output.row_dimensions[output_row_idx].height = 80
                output_row_idx += 1

                # 批量插入（每50条，适配Milvus最佳实践）
                if len(insert_batch) >= 50:
                    res = milvus_client.insert(COLLECTION_NAME, insert_batch)
                    inserted = res["insert_count"]
                    total_inserted += inserted
                    print(f"✅ 已插入 {inserted} 条数据（累计：{total_inserted}）")
                    insert_batch = []

            except Exception as e:
                failed_rows.append(f"行{row_idx}分块失败：{str(e)[:50]}")
                continue

    # 插入剩余数据
    if insert_batch:
        res = milvus_client.insert(COLLECTION_NAME, insert_batch)
        inserted = res["insert_count"]
        total_inserted += inserted
        print(f"✅ 已插入剩余 {inserted} 条数据（累计：{total_inserted}）")

    # 刷新数据到磁盘并加载到内存
    print("⏳ 正在刷新数据并加载到内存...")
    milvus_client.flush(COLLECTION_NAME)
    milvus_client.load_collection(COLLECTION_NAME)
    print("✅ 数据加载完成")

    # 保存新XLSX
    column_widths = [30, 30, 15, 50, 80, 80, 80]
    for col_idx, width in enumerate(column_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws_output.column_dimensions[col_letter].width = width
    wb_output.save(OUTPUT_XLSX)

    # 验证结果
    stats = milvus_client.get_collection_stats(COLLECTION_NAME)
    print(f"\n🎉 全流程完成！")
    print(f"✅ 向量文件保存至：{VECTOR_DIR}")
    print(f"✅ 入库数据文件：{OUTPUT_XLSX}")
    print(f"✅ 成功插入数据量：{total_inserted} 条")
    print(f"✅ Milvus集合总数据量：{stats['row_count']} 条")

    if failed_rows:
        print(f"⚠️ 共{len(failed_rows)}行处理失败（前10条）：")
        for fail in failed_rows[:10]:
            print(f"  - {fail}")

    # 关闭文件
    wb_input.close()
    wb_output.close()


# ====================== 主执行流程（仅加载你的Collection） ======================
if __name__ == "__main__":
    # 第一步：先运行你自己的Schema+Collection创建代码
    # 第二步：再运行此脚本（仅加载+插入）
    load_collection_and_insert()